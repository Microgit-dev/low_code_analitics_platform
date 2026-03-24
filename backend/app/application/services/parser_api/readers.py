from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook

from .models import TableSheet


class TableReader:
    def read(self, path: str | Path, sheet_name: str | None = None) -> list[TableSheet]:
        path = Path(path)
        suffix = path.suffix.lower()
        if suffix == ".csv":
            return self._read_csv(path)
        if suffix == ".xls":
            return self._read_xls(path, sheet_name)
        if suffix == ".xlsx":
            return self._read_xlsx(path, sheet_name)
        raise ValueError(f"Unsupported table format: {suffix}")

    def _read_csv(self, path: Path) -> list[TableSheet]:
        with path.open("r", encoding="utf-8-sig", newline="") as file:
            rows = [tuple(row) for row in csv.reader(file)]

        nrows = len(rows)
        ncols = max((len(row) for row in rows), default=0)
        cells = tuple(
            tuple(row[column_index] if column_index < len(row) else "" for column_index in range(ncols))
            for row in rows
        )
        return [
            TableSheet(
                name=path.stem,
                nrows=nrows,
                ncols=ncols,
                merged_cells=(),
                cells=cells,
                source_format="csv",
            )
        ]

    def _read_xls(self, path: Path, sheet_name: str | None) -> list[TableSheet]:
        try:
            import xlrd
        except ModuleNotFoundError as exc:
            raise ModuleNotFoundError(
                "xlrd is required to read .xls files. Install dependency 'xlrd'."
            ) from exc

        workbook = xlrd.open_workbook(str(path), formatting_info=True)
        sheets = [workbook.sheet_by_name(sheet_name)] if sheet_name else workbook.sheets()
        result: list[TableSheet] = []
        for raw_sheet in sheets:
            if not raw_sheet.nrows or not raw_sheet.ncols:
                continue
            cells = tuple(
                tuple(raw_sheet.cell_value(row_index, column_index) for column_index in range(raw_sheet.ncols))
                for row_index in range(raw_sheet.nrows)
            )
            result.append(
                TableSheet(
                    name=raw_sheet.name,
                    nrows=raw_sheet.nrows,
                    ncols=raw_sheet.ncols,
                    merged_cells=tuple(raw_sheet.merged_cells),
                    cells=cells,
                    source_format="xls",
                )
            )
        return result

    def _read_xlsx(self, path: Path, sheet_name: str | None) -> list[TableSheet]:
        workbook = load_workbook(path, data_only=True)
        names = [sheet_name] if sheet_name else workbook.sheetnames
        result: list[TableSheet] = []
        for name in names:
            worksheet = workbook[name]
            nrows = worksheet.max_row
            ncols = worksheet.max_column
            if not nrows or not ncols:
                continue
            cells = tuple(
                tuple(
                    worksheet.cell(row=row_index + 1, column=column_index + 1).value
                    for column_index in range(ncols)
                )
                for row_index in range(nrows)
            )
            merged_cells = tuple(
                (cell_range.min_row - 1, cell_range.max_row, cell_range.min_col - 1, cell_range.max_col)
                for cell_range in worksheet.merged_cells.ranges
            )
            result.append(
                TableSheet(
                    name=name,
                    nrows=nrows,
                    ncols=ncols,
                    merged_cells=merged_cells,
                    cells=cells,
                    source_format="xlsx",
                )
            )
        return result
